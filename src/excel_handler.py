import os
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


class ExcelReader:
    """Read company inputs from Excel workbooks."""

    NAME_HEADER_KEYWORDS = ("company name", "english")
    TAX_CODE_HEADER_KEYWORDS = ("tax code",)
    MAX_HEADER_SCAN_ROWS = 20

    # -- Public API --
    def read_companies(self, file_path: str) -> List[Dict]:
        """Read company names and tax codes from an Excel file.

        Args:
            file_path: Path to the Excel workbook.

        Returns:
            List of dictionaries with original_name and tax_code keys.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the workbook is empty or does not contain expected headers.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as error:
            raise ValueError(f"Invalid Excel file: {file_path}") from error

        try:
            worksheet = workbook.active
            if worksheet.max_row == 0 or worksheet.max_column == 0:
                raise ValueError("Excel sheet is empty.")

            header_row, name_col, tax_code_col = self._detect_columns(worksheet)
            companies: List[Dict] = []

            for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
                original_name = self._clean_text(self._get_cell_value(row, name_col))
                if not original_name:
                    continue

                tax_code = ""
                if tax_code_col is not None:
                    tax_code = self._clean_optional_text(
                        self._get_cell_value(row, tax_code_col)
                    )

                companies.append({
                    "original_name": original_name,
                    "tax_code": tax_code,
                })

            return companies
        finally:
            workbook.close()

    # -- Private helpers --
    def _detect_columns(self, worksheet: Any) -> Tuple[int, int, Optional[int]]:
        has_content = False
        max_row = worksheet.max_row or self.MAX_HEADER_SCAN_ROWS

        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(max_row, self.MAX_HEADER_SCAN_ROWS),
                values_only=True,
            ),
            start=1,
        ):
            if any(value is not None for value in row):
                has_content = True

            normalized_headers = [self._normalize_header(value) for value in row]
            name_col = self._find_header_index(
                normalized_headers,
                self.NAME_HEADER_KEYWORDS,
            )
            if name_col is None:
                continue

            tax_code_col = self._find_header_index(
                normalized_headers,
                self.TAX_CODE_HEADER_KEYWORDS,
            )
            return row_index, name_col, tax_code_col

        if not has_content:
            raise ValueError("Excel sheet is empty.")

        raise ValueError(
            "Invalid Excel format: missing company name or english header."
        )

    def _find_header_index(
        self,
        headers: List[str],
        keywords: Tuple[str, ...],
    ) -> Optional[int]:
        for index, header in enumerate(headers):
            if any(keyword in header for keyword in keywords):
                return index
        return None

    def _normalize_header(self, value: Any) -> str:
        if value is None:
            return ""
        return " ".join(str(value).strip().lower().split())

    def _get_cell_value(self, row: tuple, index: int) -> Any:
        if index >= len(row):
            return None
        return row[index]

    def _clean_text(self, value: Any) -> str:
        if not isinstance(value, str):
            return ""
        return value.strip()

    def _clean_optional_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()


class ExcelWriter:
    """Write pipeline results and final reports to Excel workbooks."""

    DETAIL_COLUMNS = [
        (("original_name", "vietnamese_name"), "Tên công ty"),
        (("tax_code",), "MST"),
        (("phone",), "Phone"),
        (("email",), "Email"),
        (("address",), "Address"),
        (("website",), "Website"),
        (("source", "source_type", "source_url"), "Source"),
        (("confidence", "confidence_score"), "Confidence"),
    ]

    BASIC_COLUMNS = [
        (("original_name", "vietnamese_name"), "Tên công ty"),
        (("tax_code",), "MST"),
        (("phone",), "Phone"),
        (("email",), "Email"),
        (("address",), "Address"),
        (("website",), "Website"),
        (("source", "source_type", "source_url"), "Source"),
        (("confidence", "confidence_score"), "Confidence"),
        (("status",), "Status"),
        (("error", "error_message"), "Error"),
    ]

    # -- Public API --
    def write_results(self, output_path: str, results: List[Dict]) -> None:
        """Write a basic one-sheet Excel report.

        Args:
            output_path: Path where the workbook should be saved.
            results: List of result dictionaries.
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Kết quả"

        self._write_table(worksheet, self.BASIC_COLUMNS, results)
        self._save_workbook(workbook, output_path)

    def write_final_report(
        self,
        output_path: str,
        aggregated_data: List[Dict],
        summary_stats: Dict,
    ) -> None:
        """Write a two-sheet final Excel report.

        Args:
            output_path: Path where the workbook should be saved.
            aggregated_data: Company-level aggregated contact data.
            summary_stats: Pipeline summary statistics.
        """
        workbook = Workbook()
        detail_sheet = workbook.active
        detail_sheet.title = "Chi tiết"

        self._write_table(detail_sheet, self.DETAIL_COLUMNS, aggregated_data)

        stats_sheet = workbook.create_sheet("Thống kê")
        self._write_summary(stats_sheet, aggregated_data, summary_stats)
        self._save_workbook(workbook, output_path)

    # -- Private helpers --
    def _write_table(
        self,
        worksheet: Any,
        columns: List[Tuple[Tuple[str, ...], str]],
        rows: List[Dict],
    ) -> None:
        worksheet.append([title for _, title in columns])
        self._style_header(worksheet)

        for row in rows:
            worksheet.append([self._get_row_value(row, keys) for keys, _ in columns])

        worksheet.freeze_panes = "A2"
        self._auto_width_columns(worksheet)

    def _write_summary(
        self,
        worksheet: Any,
        aggregated_data: List[Dict],
        summary_stats: Dict,
    ) -> None:
        worksheet.append(["Chỉ số", "Giá trị"])
        self._style_header(worksheet)

        total_companies = summary_stats.get("total_companies", len(aggregated_data))
        phone_found_percent = summary_stats.get(
            "phone_found_percent",
            self._calculate_phone_found_percent(aggregated_data),
        )
        credits_used = summary_stats.get("credits_used", 0)

        worksheet.append(["Tổng công ty", total_companies])
        worksheet.append(["% tìm được phone", phone_found_percent])

        step_percentages = self._get_step_percentages(summary_stats)
        for step_name, percentage in step_percentages.items():
            worksheet.append([f"% {step_name}", percentage])

        worksheet.append(["Credits used", credits_used])
        worksheet.freeze_panes = "A2"
        self._auto_width_columns(worksheet)

    def _get_step_percentages(self, summary_stats: Dict) -> Dict:
        for key in ("step_percentages", "percent_by_step", "steps"):
            value = summary_stats.get(key)
            if isinstance(value, dict):
                return value
        return {}

    def _calculate_phone_found_percent(self, aggregated_data: List[Dict]) -> float:
        if not aggregated_data:
            return 0.0
        found_count = sum(1 for item in aggregated_data if item.get("phone"))
        return round(found_count / len(aggregated_data) * 100, 2)

    def _get_row_value(self, row: Dict, keys: Tuple[str, ...]) -> Any:
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                return value
        return ""

    def _style_header(self, worksheet: Any) -> None:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    def _auto_width_columns(self, worksheet: Any) -> None:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))

            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                60,
            )

    def _save_workbook(self, workbook: Workbook, output_path: str) -> None:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        workbook.save(output_path)
